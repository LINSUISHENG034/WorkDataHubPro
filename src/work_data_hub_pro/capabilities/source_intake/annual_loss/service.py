from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, datetime
from hashlib import sha256
from pathlib import Path

from openpyxl import load_workbook

from work_data_hub_pro.platform.contracts.models import (
    FieldTraceEvent,
    InputBatch,
    InputRecord,
)


@dataclass(frozen=True)
class IntakeResult:
    batch: InputBatch
    records: list[InputRecord]
    trace_events: list[FieldTraceEvent]


class AnnualLossIntakeService:
    domain = "annual_loss"
    sheet_names = ("企年受托流失(解约)", "企年投资流失(解约)")
    stage_id = "source_intake"

    def read_batch(
        self,
        *,
        run_id: str,
        period: str,
        source_files: list[Path],
    ) -> IntakeResult:
        snapshot_hash = sha256()
        records: list[InputRecord] = []
        trace_events: list[FieldTraceEvent] = []
        merged_anchor_row_no = 2

        for source_file in source_files:
            workbook = load_workbook(source_file, read_only=True, data_only=True)
            try:
                for sheet_name in self.sheet_names:
                    sheet = workbook[sheet_name]
                    headers = [
                        cell.value
                        for cell in next(sheet.iter_rows(min_row=1, max_row=1))
                    ]
                    for source_row_no, row in enumerate(
                        sheet.iter_rows(min_row=2, values_only=True),
                        start=2,
                    ):
                        if all(
                            value is None
                            or (isinstance(value, str) and value.strip() == "")
                            for value in row
                        ):
                            continue
                        payload = dict(zip(headers, row, strict=True)) | {
                            "source_sheet": sheet_name,
                            "source_row_no": source_row_no,
                        }
                        snapshot_hash.update(
                            repr(
                                (
                                    source_file.name,
                                    sheet_name,
                                    source_row_no,
                                    payload,
                                )
                            ).encode("utf-8")
                        )
                        record_id = (
                            f"{run_id}:{source_file.stem}:{sheet_name}:{source_row_no}"
                        )
                        record = InputRecord(
                            run_id=run_id,
                            record_id=record_id,
                            batch_id=f"{self.domain}:{period}",
                            anchor_row_no=merged_anchor_row_no,
                            origin_row_nos=[source_row_no],
                            parent_record_ids=[],
                            stage_row_no=merged_anchor_row_no,
                            raw_payload=payload,
                        )
                        records.append(record)
                        trace_events.append(
                            FieldTraceEvent(
                                trace_id=f"trace:{record_id}",
                                event_id=f"{record_id}:intake",
                                event_seq=0,
                                run_id=run_id,
                                batch_id=record.batch_id,
                                record_id=record.record_id,
                                anchor_row_no=record.anchor_row_no,
                                stage_id=self.stage_id,
                                field_name="raw_payload",
                                value_before=None,
                                value_after=payload,
                                rule_id="capture-input",
                                rule_version="1",
                                config_release_id="system:source-intake",
                                action_type="snapshot",
                                timestamp=datetime.now(UTC).isoformat(),
                                success=True,
                            )
                        )
                        merged_anchor_row_no += 1
            finally:
                workbook.close()

        batch = InputBatch(
            batch_id=f"{self.domain}:{period}",
            domain=self.domain,
            period=period,
            source_files=[str(path) for path in source_files],
            input_snapshot_id=snapshot_hash.hexdigest(),
            row_count=len(records),
        )
        return IntakeResult(batch=batch, records=records, trace_events=trace_events)
