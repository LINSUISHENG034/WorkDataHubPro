from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, datetime
from hashlib import sha256
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from work_data_hub_pro.platform.contracts.models import (
    FieldTraceEvent,
    InputBatch,
    InputRecord,
)
from work_data_hub_pro.platform.contracts.validators import (
    validate_input_batch,
    validate_input_record,
)


@dataclass(frozen=True)
class IntakeResult:
    batch: InputBatch
    records: list[InputRecord]
    trace_events: list[FieldTraceEvent]


class AnnuityPerformanceIntakeService:
    domain = "annuity_performance"
    sheet_name = "AnnuityPerformance"
    stage_id = "source_intake"
    _field_aliases = {
        "月度": "period",
        "period": "period",
        "业务类型": "business_type",
        "business_type": "business_type",
        "客户名称": "company_name",
        "客户全称": "company_name",
        "company_name": "company_name",
        "计划代码": "plan_code",
        "plan_code": "plan_code",
        "计划类型": "plan_type",
        "plan_type": "plan_type",
        "期末资产规模": "ending_assets",
        "ending_assets": "ending_assets",
        "sales_amount": "ending_assets",
        "公司代码": "source_company_id",
        "集团企业客户号": "source_company_id",
        "source_company_id": "source_company_id",
    }
    _non_golden_optional_fields = (
        "plan_code",
        "plan_type",
        "source_company_id",
    )

    def read_batch(
        self,
        *,
        run_id: str,
        period: str,
        source_files: list[Path],
    ) -> IntakeResult:
        snapshot_hash = sha256()
        records: list[InputRecord] = []
        trace_events: list[FieldTraceEvent] = []

        for source_file in source_files:
            workbook = load_workbook(source_file, read_only=True, data_only=True)
            try:
                sheet = workbook[self.sheet_name]
                headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

                for row_no, row in enumerate(
                    sheet.iter_rows(min_row=2, values_only=True),
                    start=2,
                ):
                    payload = self._normalize_payload(
                        headers=headers,
                        row=row,
                        source_file=source_file,
                        row_no=row_no,
                    )
                    snapshot_hash.update(
                        repr((source_file.name, row_no, payload)).encode("utf-8")
                    )
                    record_id = f"{run_id}:{source_file.stem}:{row_no}"
                    record = InputRecord(
                        run_id=run_id,
                        record_id=record_id,
                        batch_id=f"{self.domain}:{period}",
                        anchor_row_no=row_no,
                        origin_row_nos=[row_no],
                        parent_record_ids=[],
                        stage_row_no=row_no,
                        raw_payload=payload,
                    )
                    validate_input_record(
                        record,
                        required_fields=(
                            "period",
                            "business_type",
                            "company_name",
                            "ending_assets",
                        ),
                        alternative_field_groups=(("plan_code", "plan_type"),),
                    )
                    records.append(record)
                    trace_events.append(
                        FieldTraceEvent(
                            trace_id=f"trace:{record_id}",
                            event_id=f"{record_id}:intake",
                            event_seq=0,
                            run_id=run_id,
                            batch_id=record.batch_id,
                            record_id=record.record_id,
                            anchor_row_no=record.anchor_row_no,
                            stage_id=self.stage_id,
                            field_name="raw_payload",
                            value_before=None,
                            value_after=payload,
                            rule_id="capture-input",
                            rule_version="1",
                            config_release_id="system:source-intake",
                            action_type="snapshot",
                            timestamp=datetime.now(UTC).isoformat(),
                            success=True,
                        )
                    )
            finally:
                workbook.close()

        batch = InputBatch(
            batch_id=f"{self.domain}:{period}",
            domain=self.domain,
            period=period,
            source_files=[str(path) for path in source_files],
            input_snapshot_id=snapshot_hash.hexdigest(),
            row_count=len(records),
        )
        validate_input_batch(batch)
        return IntakeResult(batch=batch, records=records, trace_events=trace_events)

    def _normalize_payload(
        self,
        *,
        headers: list[object],
        row: tuple[object, ...],
        source_file: Path,
        row_no: int,
    ) -> dict[str, Any]:
        aliases_applied: dict[str, str] = {}
        ignored_columns: list[str] = []
        normalized: dict[str, Any] = {
            "source_file": source_file.name,
            "source_row_no": row_no,
        }

        for header, value in zip(headers, row, strict=True):
            header_name = str(header)
            canonical_name = self._field_aliases.get(header_name)
            if canonical_name is None:
                ignored_columns.append(header_name)
                continue
            if header_name != canonical_name:
                aliases_applied[header_name] = canonical_name
            normalized[canonical_name] = value

        derived_fields: dict[str, Any] = {}
        if "business_type" not in normalized:
            normalized["business_type"] = self.domain
            derived_fields["business_type"] = self.domain
        if "ending_assets" in normalized and "sales_amount" not in normalized:
            normalized["sales_amount"] = normalized["ending_assets"]
            derived_fields["sales_amount"] = normalized["ending_assets"]

        normalized["source_intake_adaptation"] = {
            "source_headers": [str(header) for header in headers],
            "aliases_applied": aliases_applied,
            "ignored_columns": ignored_columns,
            "missing_non_golden_columns": [
                field_name
                for field_name in self._non_golden_optional_fields
                if field_name not in normalized
            ],
            "derived_fields": derived_fields,
        }
        return normalized
