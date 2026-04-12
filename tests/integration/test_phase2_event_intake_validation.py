from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from work_data_hub_pro.capabilities.source_intake.annual_award.service import (
    AnnualAwardIntakeService,
)
from work_data_hub_pro.capabilities.source_intake.annual_loss.service import (
    AnnualLossIntakeService,
)


def _write_award_workbook(path: Path) -> None:
    workbook = Workbook()
    trustee = workbook.active
    trustee.title = "TrusteeAwards"
    trustee.append(["上报月份", "客户全称", "计划类型", "奖励金额", "附加列"])
    trustee.append(["2026-03", "Acme", "集合", "5000", "drop-me"])
    investee = workbook.create_sheet("InvesteeAwards")
    investee.append(["上报月份", "客户全称", "计划类型", "奖励金额"])
    investee.append(["2026-03", "Beta", "单一", "1000"])
    workbook.save(path)


def _write_loss_workbook(path: Path) -> None:
    workbook = Workbook()
    trustee = workbook.active
    trustee.title = "企年受托流失(解约)"
    trustee.append(
        [
            "上报月份",
            "业务类型",
            "客户全称",
            "计划类型",
            "年金计划号",
            "流失日期",
            "附加列",
        ]
    )
    trustee.append(["2026-03", "受托", "共享客户（流失）", "集合", "", "2026-03-15", "drop-me"])
    investee = workbook.create_sheet("企年投资流失(解约)")
    investee.append(
        [
            "上报月份",
            "业务类型",
            "客户全称",
            "计划类型",
            "年金计划号",
            "流失日期",
        ]
    )
    investee.append(["2026-03", "投管", "新客流失", "单一", "", "2026-03-15"])
    workbook.save(path)


def test_event_domain_intake_accepts_aliases_and_records_adaptation(tmp_path) -> None:
    award_workbook = tmp_path / "annual_award_real_style.xlsx"
    loss_workbook = tmp_path / "annual_loss_real_style.xlsx"
    _write_award_workbook(award_workbook)
    _write_loss_workbook(loss_workbook)

    award_result = AnnualAwardIntakeService().read_batch(
        run_id="run-001",
        period="2026-03",
        source_files=[award_workbook],
    )
    loss_result = AnnualLossIntakeService().read_batch(
        run_id="run-001",
        period="2026-03",
        source_files=[loss_workbook],
    )

    award_record = award_result.records[0]
    award_adaptation = award_record.raw_payload["source_intake_adaptation"]
    assert award_record.raw_payload["company_name"] == "Acme"
    assert award_record.raw_payload["period"] == "2026-03"
    assert award_record.raw_payload["source_sheet"] == "TrusteeAwards"
    assert award_adaptation["aliases_applied"]["客户全称"] == "company_name"
    assert award_adaptation["ignored_columns"] == ["附加列"]
    assert "plan_code" in award_adaptation["missing_non_golden_columns"]

    loss_record = loss_result.records[0]
    loss_adaptation = loss_record.raw_payload["source_intake_adaptation"]
    assert loss_record.raw_payload["company_name"] == "共享客户（流失）"
    assert loss_record.raw_payload["period"] == "2026-03"
    assert loss_record.raw_payload["source_sheet"] == "企年受托流失(解约)"
    assert loss_adaptation["ignored_columns"] == ["附加列"]
    assert "plan_code" in loss_adaptation["missing_non_golden_columns"]


@pytest.mark.parametrize(
    ("service_cls", "builder"),
    [
        (AnnualAwardIntakeService, _write_award_workbook),
        (AnnualLossIntakeService, _write_loss_workbook),
    ],
)
def test_event_domain_intake_rejects_missing_minimum_skeleton(
    tmp_path,
    service_cls,
    builder,
) -> None:
    workbook_path = tmp_path / f"{service_cls.__name__}.xlsx"
    builder(workbook_path)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TrusteeAwards" if service_cls is AnnualAwardIntakeService else "企年受托流失(解约)"
    sheet.append(["上报月份", "客户全称"])
    sheet.append(["2026-03", "Acme"])
    if service_cls is AnnualAwardIntakeService:
        investee = workbook.create_sheet("InvesteeAwards")
        investee.append(["上报月份", "客户全称"])
        investee.append(["2026-03", "Beta"])
    else:
        investee = workbook.create_sheet("企年投资流失(解约)")
        investee.append(["上报月份", "客户全称"])
        investee.append(["2026-03", "Beta"])
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="plan_code or plan_type"):
        service_cls().read_batch(
            run_id="run-001",
            period="2026-03",
            source_files=[workbook_path],
        )
