from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

import work_data_hub_pro.apps.orchestration.replay.runtime as replay_runtime
from work_data_hub_pro.apps.orchestration.replay.annual_award_slice import (
    run_annual_award_slice,
)
from work_data_hub_pro.apps.orchestration.replay.annual_loss_slice import (
    run_annual_loss_slice,
)
from work_data_hub_pro.apps.orchestration.replay.annuity_performance_slice import (
    run_annuity_performance_slice,
)
from work_data_hub_pro.apps.orchestration.replay.errors import (
    ReplayAssetNotFoundError,
    ReplayConfigurationError,
    ReplayContractSetupError,
    ReplayDiagnosticsNotFoundError,
    ReplaySetupError,
    translate_replay_setup_error,
)
from work_data_hub_pro.governance.compatibility.gate_runtime import (
    load_required_checkpoint_baseline,
)
from work_data_hub_pro.platform.contracts.publication import (
    PublicationMode,
    PublicationPlan,
)
from work_data_hub_pro.platform.contracts.validators import validate_publication_plan
from work_data_hub_pro.platform.publication.service import (
    PublicationPolicy,
    PublicationPolicyEntry,
    build_publication_plan,
    load_publication_policy,
)


def _write_annuity_workbook(workbook_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "AnnuityPerformance"
    sheet.append(["company_name", "plan_code", "period", "sales_amount"])
    sheet.append(["Acme", "PLAN-A", "2026-03", "1200.50"])
    workbook.save(workbook_path)


def _write_annuity_replay_assets(
    replay_root: Path,
    *,
    include_intermediate_baselines: bool,
) -> None:
    replay_root.mkdir(parents=True, exist_ok=True)
    (replay_root / "annual_award_fixture_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "company_id": "company-001",
                    "plan_code": "PLAN-A",
                    "period": "2026-03",
                    "award_code": "AWARD-01",
                    "source_sheet": "AwardRegister",
                    "source_record_id": "award-001",
                }
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    (replay_root / "annual_loss_fixture_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "company_id": "company-001",
                    "plan_code": "PLAN-Z",
                    "period": "2026-03",
                    "loss_code": "LOSS-99",
                    "source_sheet": "LossRegister",
                    "source_record_id": "loss-001",
                }
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    (replay_root / "legacy_monthly_snapshot_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "period": "2026-03",
                    "contract_state_rows": 1,
                    "award_fixture_rows": 1,
                    "loss_fixture_rows": 0,
                }
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    if not include_intermediate_baselines:
        return

    for checkpoint_name in (
        "reference_derivation",
        "fact_processing",
        "identity_resolution",
        "contract_state",
    ):
        (replay_root / f"legacy_{checkpoint_name}_2026_03.json").write_text(
            "[]",
            encoding="utf-8",
        )


def _write_award_workbook(workbook_path: Path) -> None:
    workbook = Workbook()
    trustee = workbook.active
    trustee.title = "TrusteeAwards"
    trustee.append(
        [
            "company_name",
            "source_company_id",
            "plan_code",
            "plan_type",
            "product_line_code",
            "period",
            "award_amount",
        ]
    )
    trustee.append(["Acme", "company-001", "", "collective", "pl-ret", "2026-03", "5000"])
    investee = workbook.create_sheet("InvesteeAwards")
    investee.append(
        [
            "company_name",
            "source_company_id",
            "plan_code",
            "plan_type",
            "product_line_code",
            "period",
            "award_amount",
        ]
    )
    investee.append(["Beta", "", "", "single", "pl-alt", "2026-03", "1000"])
    workbook.save(workbook_path)


def _write_award_replay_assets(
    replay_root: Path,
    *,
    include_plan_history: bool = True,
    include_annuity_fixture: bool = True,
    include_loss_fixture: bool = True,
    include_intermediate_baselines: bool = False,
) -> None:
    replay_root.mkdir(parents=True, exist_ok=True)
    if include_annuity_fixture:
        (replay_root / "annuity_performance_fixture_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "record_id": "perf-001",
                        "company_id": "company-001",
                        "plan_code": "P9001",
                        "period": "2026-03",
                        "source_record_id": "perf-001",
                    }
                ],
                indent=2,
            ),
            encoding="utf-8",
        )
    if include_loss_fixture:
        (replay_root / "annual_loss_fixture_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "company_id": "company-001",
                        "plan_code": "LOSS-99",
                        "period": "2026-03",
                        "loss_code": "LOSS-99",
                        "source_sheet": "LossRegister",
                        "source_record_id": "loss-001",
                    }
                ],
                indent=2,
            ),
            encoding="utf-8",
        )
    if include_plan_history:
        (replay_root / "customer_plan_history_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "company_id": "company-001",
                        "product_line_code": "PL-RET",
                        "plan_code": "P9001",
                        "effective_period": "2025-12",
                    }
                ],
                indent=2,
            ),
            encoding="utf-8",
        )
    (replay_root / "legacy_monthly_snapshot_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "period": "2026-03",
                    "contract_state_rows": 1,
                    "award_fixture_rows": 1,
                    "loss_fixture_rows": 0,
                }
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    if include_intermediate_baselines:
        for checkpoint_name in (
            "reference_derivation",
            "fact_processing",
            "identity_resolution",
            "contract_state",
        ):
            (replay_root / f"legacy_{checkpoint_name}_2026_03.json").write_text(
                "[]",
                encoding="utf-8",
            )


def _write_loss_workbook(workbook_path: Path) -> None:
    workbook = Workbook()
    trustee = workbook.active
    trustee.title = "企年受托流失(解约)"
    trustee.append(
        [
            "上报月份",
            "业务类型",
            "计划类型",
            "客户全称",
            "机构",
            "年金计划号",
            "流失日期",
            "计划规模",
            "受托人",
            "company_id",
        ]
    )
    trustee.append(
        [
            "2026年03月",
            "受托",
            "集合",
            "共享客户（流失）",
            "北京",
            "",
            "",
            "80",
            "原受托机构A",
            "company-001",
        ]
    )
    investee = workbook.create_sheet("企年投资流失(解约)")
    investee.append(
        [
            "上报月份",
            "业务类型",
            "计划类型",
            "客户全称",
            "机构",
            "年金计划号",
            "流失日期",
            "计划规模",
            "受托人",
            "company_id",
        ]
    )
    investee.append(
        [
            "2026年03月",
            "投管",
            "单一",
            "新客流失",
            "未知机构",
            "",
            "2026-03-15",
            "60",
            "原受托机构B",
            "",
        ]
    )
    workbook.save(workbook_path)


def _write_loss_replay_assets(
    replay_root: Path,
    *,
    include_plan_history: bool = True,
    include_annuity_fixture: bool = True,
    include_award_fixture: bool = True,
    include_intermediate_baselines: bool = False,
) -> None:
    replay_root.mkdir(parents=True, exist_ok=True)
    if include_annuity_fixture:
        (replay_root / "annuity_performance_fixture_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "record_id": "perf-001",
                        "company_id": "company-001",
                        "plan_code": "P9001",
                        "period": "2026-03",
                        "source_record_id": "perf-001",
                    }
                ],
                indent=2,
            ),
            encoding="utf-8",
        )
    if include_award_fixture:
        (replay_root / "annual_award_fixture_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "company_id": "company-001",
                        "plan_code": "P9001",
                        "period": "2026-03",
                        "award_code": "AWARD-01",
                        "source_sheet": "AwardRegister",
                        "source_record_id": "award-001",
                    }
                ],
                indent=2,
            ),
            encoding="utf-8",
        )
    if include_plan_history:
        (replay_root / "customer_plan_history_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "company_id": "company-001",
                        "product_line_code": "PL202",
                        "plan_code": "P9001",
                        "effective_period": "2025-12",
                    }
                ],
                indent=2,
            ),
            encoding="utf-8",
        )
    (replay_root / "legacy_monthly_snapshot_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "period": "2026-03",
                    "contract_state_rows": 1,
                    "award_fixture_rows": 1,
                    "loss_fixture_rows": 1,
                }
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    if include_intermediate_baselines:
        for checkpoint_name in (
            "reference_derivation",
            "fact_processing",
            "identity_resolution",
            "contract_state",
        ):
            (replay_root / f"legacy_{checkpoint_name}_2026_03.json").write_text(
                "[]",
                encoding="utf-8",
            )


def test_translate_replay_setup_error_maps_missing_baseline() -> None:
    baseline_path = Path("reference/historical_replays/annual_award/missing.json")

    with pytest.raises(FileNotFoundError) as excinfo:
        load_required_checkpoint_baseline(baseline_path, "reference_derivation")

    error = translate_replay_setup_error(
        domain="annual_award",
        stage="baseline_load",
        exc=excinfo.value,
        context={"baseline_path": str(baseline_path)},
    )

    assert isinstance(error, ReplayAssetNotFoundError)
    assert isinstance(error, ReplaySetupError)
    assert error.domain == "annual_award"
    assert error.stage == "baseline_load"
    assert error.original_exception_type == "FileNotFoundError"
    assert "Missing accepted baseline" in error.original_exception_message
    assert error.context["baseline_path"] == str(baseline_path)


def test_translate_replay_setup_error_maps_malformed_baseline_json_type(
    tmp_path: Path,
) -> None:
    baseline_path = tmp_path / "legacy_identity_resolution_2026_03.json"
    baseline_path.write_text(
        json.dumps({"not": "a list"}),
        encoding="utf-8",
    )

    with pytest.raises(TypeError) as excinfo:
        load_required_checkpoint_baseline(baseline_path, "identity_resolution")

    error = translate_replay_setup_error(
        domain="annual_loss",
        stage="baseline_load",
        exc=excinfo.value,
        context={"baseline_path": str(baseline_path)},
    )

    assert isinstance(error, ReplayConfigurationError)
    assert error.original_exception_type == "TypeError"
    assert "must contain a JSON array" in error.original_exception_message


@pytest.mark.parametrize(
    ("stage", "exc"),
    [
        (
            "publication_policy_domain",
            KeyError("annual_award"),
        ),
        (
            "publication_policy_target",
            KeyError("monthly_snapshot"),
        ),
    ],
)
def test_translate_replay_setup_error_maps_missing_publication_policy_domain_or_target(
    stage: str,
    exc: KeyError,
) -> None:
    error = translate_replay_setup_error(
        domain="annuity_performance",
        stage=stage,
        exc=exc,
        context={"policy_path": "config/policies/publication.json"},
    )

    assert isinstance(error, ReplayConfigurationError)
    assert error.original_exception_type == "KeyError"
    assert error.context["policy_path"] == "config/policies/publication.json"


def test_missing_publication_policy_domain_translates_from_real_loader(
    tmp_path: Path,
) -> None:
    policy_path = tmp_path / "publication.json"
    policy_path.write_text(
        json.dumps(
            {
                "annual_loss": {
                    "monthly_snapshot": {
                        "mode": "REFRESH",
                        "transaction_group": "monthly-snapshot",
                        "idempotency_scope": "period",
                    }
                }
            }
        ),
        encoding="utf-8",
    )

    with pytest.raises(KeyError) as excinfo:
        load_publication_policy(policy_path, domain="annual_award")

    error = translate_replay_setup_error(
        domain="annual_award",
        stage="publication_policy_domain",
        exc=excinfo.value,
        context={"policy_path": str(policy_path)},
    )

    assert isinstance(error, ReplayConfigurationError)
    assert error.original_exception_message == "'annual_award'"


def test_missing_publication_policy_target_translates_from_real_builder() -> None:
    policy = PublicationPolicy(
        domain="annual_award",
        targets={
            "fact_annuity_performance": PublicationPolicyEntry(
                mode=PublicationMode.REFRESH,
                transaction_group="facts",
                idempotency_scope="batch",
            )
        },
    )

    with pytest.raises(KeyError) as excinfo:
        build_publication_plan(
            policy=policy,
            publication_id="publication-monthly-snapshot",
            target_name="monthly_snapshot",
            target_kind="projection",
            refresh_keys=["period"],
            upsert_keys=[],
            source_batch_id="annual_award:2026-03",
            source_run_id="run-001",
        )

    error = translate_replay_setup_error(
        domain="annual_award",
        stage="publication_policy_target",
        exc=excinfo.value,
        context={"target_name": "monthly_snapshot"},
    )

    assert isinstance(error, ReplayConfigurationError)
    assert error.original_exception_message == "'monthly_snapshot'"


def test_translate_replay_setup_error_maps_publication_plan_validation_failures() -> None:
    plan = PublicationPlan(
        publication_id="publication-facts",
        target_name="fact_annuity_performance",
        target_kind="fact",
        mode=PublicationMode.REFRESH,
        refresh_keys=[],
        upsert_keys=[],
        source_batch_id="annuity_performance:2026-03",
        source_run_id="run-001",
        idempotency_scope="batch",
        transaction_group="facts",
    )

    with pytest.raises(ValueError) as excinfo:
        validate_publication_plan(plan)

    error = translate_replay_setup_error(
        domain="annuity_performance",
        stage="publication_plan_validation",
        exc=excinfo.value,
        context={"target_name": plan.target_name},
    )

    assert isinstance(error, ReplayContractSetupError)
    assert error.original_exception_type == "ValueError"
    assert "REFRESH publication plans must define refresh_keys." in (
        error.original_exception_message
    )


def test_replay_diagnostics_not_found_error_is_typed_setup_error() -> None:
    error = ReplayDiagnosticsNotFoundError(
        domain="annual_award",
        stage="diagnostics_lookup",
        message="Replay diagnostics package was not found.",
        context={"comparison_run_id": "missing-run"},
        original_exception_type="FileNotFoundError",
        original_exception_message="comparison_run_id not found: missing-run",
    )

    assert isinstance(error, ReplaySetupError)
    assert error.context["comparison_run_id"] == "missing-run"


def test_annuity_runner_raises_typed_missing_baseline(tmp_path: Path) -> None:
    workbook_path = tmp_path / "annuity_performance_2026_03.xlsx"
    replay_root = tmp_path / "reference" / "historical_replays" / "annuity_performance"
    _write_annuity_workbook(workbook_path)
    _write_annuity_replay_assets(
        replay_root,
        include_intermediate_baselines=False,
    )

    with pytest.raises(ReplayAssetNotFoundError) as excinfo:
        run_annuity_performance_slice(
            workbook=workbook_path,
            period="2026-03",
            replay_root=replay_root,
        )

    assert excinfo.value.stage == "baseline_load"
    assert excinfo.value.context["checkpoint_name"] == "reference_derivation"


def test_annuity_runner_raises_typed_publication_policy_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook_path = tmp_path / "annuity_performance_2026_03.xlsx"
    replay_root = tmp_path / "reference" / "historical_replays" / "annuity_performance"
    _write_annuity_workbook(workbook_path)
    _write_annuity_replay_assets(
        replay_root,
        include_intermediate_baselines=False,
    )

    def _raise_missing_domain(*args, **kwargs):
        raise KeyError("annuity_performance")

    monkeypatch.setattr(
        "work_data_hub_pro.apps.orchestration.replay.annuity_performance_slice.load_publication_policy",
        _raise_missing_domain,
    )

    with pytest.raises(ReplayConfigurationError) as excinfo:
        run_annuity_performance_slice(
            workbook=workbook_path,
            period="2026-03",
            replay_root=replay_root,
        )

    assert excinfo.value.stage == "publication_policy_domain"


def test_annuity_runner_raises_typed_publication_plan_validation_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook_path = tmp_path / "annuity_performance_2026_03.xlsx"
    replay_root = tmp_path / "reference" / "historical_replays" / "annuity_performance"
    _write_annuity_workbook(workbook_path)
    _write_annuity_replay_assets(
        replay_root,
        include_intermediate_baselines=True,
    )

    def _raise_invalid_plan(*args, **kwargs) -> None:
        raise ValueError("invalid publication plan")

    monkeypatch.setattr(replay_runtime, "validate_publication_plan", _raise_invalid_plan)

    with pytest.raises(ReplayContractSetupError) as excinfo:
        run_annuity_performance_slice(
            workbook=workbook_path,
            period="2026-03",
            replay_root=replay_root,
        )

    assert excinfo.value.stage == "publication_plan_validation"


def test_annual_award_runner_raises_typed_missing_plan_history(tmp_path: Path) -> None:
    workbook_path = tmp_path / "annual_award_2026_03.xlsx"
    replay_root = tmp_path / "reference" / "historical_replays" / "annual_award"
    _write_award_workbook(workbook_path)
    _write_award_replay_assets(
        replay_root,
        include_plan_history=False,
        include_annuity_fixture=True,
        include_loss_fixture=True,
        include_intermediate_baselines=False,
    )

    with pytest.raises(ReplayAssetNotFoundError) as excinfo:
        run_annual_award_slice(
            workbook=workbook_path,
            period="2026-03",
            replay_root=replay_root,
        )

    assert excinfo.value.stage == "plan_history_load"


def test_annual_award_runner_raises_typed_missing_fixture(tmp_path: Path) -> None:
    workbook_path = tmp_path / "annual_award_2026_03.xlsx"
    replay_root = tmp_path / "reference" / "historical_replays" / "annual_award"
    _write_award_workbook(workbook_path)
    _write_award_replay_assets(
        replay_root,
        include_plan_history=True,
        include_annuity_fixture=True,
        include_loss_fixture=False,
        include_intermediate_baselines=True,
    )

    with pytest.raises(ReplayAssetNotFoundError) as excinfo:
        run_annual_award_slice(
            workbook=workbook_path,
            period="2026-03",
            replay_root=replay_root,
        )

    assert excinfo.value.stage == "fixture_load"


def test_annual_award_runner_raises_typed_publication_policy_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook_path = tmp_path / "annual_award_2026_03.xlsx"
    replay_root = tmp_path / "reference" / "historical_replays" / "annual_award"
    _write_award_workbook(workbook_path)
    _write_award_replay_assets(
        replay_root,
        include_plan_history=True,
        include_annuity_fixture=True,
        include_loss_fixture=True,
        include_intermediate_baselines=True,
    )

    def _raise_missing_domain(*args, **kwargs):
        raise KeyError("annual_award")

    monkeypatch.setattr(
        "work_data_hub_pro.apps.orchestration.replay.annual_award_slice.load_publication_policy",
        _raise_missing_domain,
    )

    with pytest.raises(ReplayConfigurationError) as excinfo:
        run_annual_award_slice(
            workbook=workbook_path,
            period="2026-03",
            replay_root=replay_root,
        )

    assert excinfo.value.stage == "publication_policy_domain"


def test_annual_loss_runner_raises_typed_missing_plan_history(tmp_path: Path) -> None:
    workbook_path = tmp_path / "annual_loss_2026_03.xlsx"
    replay_root = tmp_path / "reference" / "historical_replays" / "annual_loss"
    _write_loss_workbook(workbook_path)
    _write_loss_replay_assets(
        replay_root,
        include_plan_history=False,
        include_annuity_fixture=True,
        include_award_fixture=True,
        include_intermediate_baselines=False,
    )

    with pytest.raises(ReplayAssetNotFoundError) as excinfo:
        run_annual_loss_slice(
            workbook=workbook_path,
            period="2026-03",
            replay_root=replay_root,
        )

    assert excinfo.value.stage == "plan_history_load"


def test_annual_loss_runner_raises_typed_missing_fixture(tmp_path: Path) -> None:
    workbook_path = tmp_path / "annual_loss_2026_03.xlsx"
    replay_root = tmp_path / "reference" / "historical_replays" / "annual_loss"
    _write_loss_workbook(workbook_path)
    _write_loss_replay_assets(
        replay_root,
        include_plan_history=True,
        include_annuity_fixture=True,
        include_award_fixture=False,
        include_intermediate_baselines=True,
    )

    with pytest.raises(ReplayAssetNotFoundError) as excinfo:
        run_annual_loss_slice(
            workbook=workbook_path,
            period="2026-03",
            replay_root=replay_root,
        )

    assert excinfo.value.stage == "fixture_load"


def test_annual_loss_runner_raises_typed_publication_plan_validation_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook_path = tmp_path / "annual_loss_2026_03.xlsx"
    replay_root = tmp_path / "reference" / "historical_replays" / "annual_loss"
    _write_loss_workbook(workbook_path)
    _write_loss_replay_assets(
        replay_root,
        include_plan_history=True,
        include_annuity_fixture=True,
        include_award_fixture=True,
        include_intermediate_baselines=True,
    )

    def _raise_invalid_plan(*args, **kwargs) -> None:
        raise ValueError("invalid publication plan")

    monkeypatch.setattr(replay_runtime, "validate_publication_plan", _raise_invalid_plan)

    with pytest.raises(ReplayContractSetupError) as excinfo:
        run_annual_loss_slice(
            workbook=workbook_path,
            period="2026-03",
            replay_root=replay_root,
        )

    assert excinfo.value.stage == "publication_plan_validation"
