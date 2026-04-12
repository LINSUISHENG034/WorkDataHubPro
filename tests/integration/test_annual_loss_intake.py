from openpyxl import Workbook

from work_data_hub_pro.capabilities.source_intake.annual_loss.service import (
    AnnualLossIntakeService,
)


def test_annual_loss_intake_merges_trustee_and_investee_sheets_into_stable_anchor_sequence(
    tmp_path,
) -> None:
    workbook_path = tmp_path / "annual_loss_2026_03.xlsx"
    workbook = Workbook()
    trustee = workbook.active
    trustee.title = "企年受托流失(解约)"
    trustee.append(
        [
            "上报月份",
            "业务类型",
            "计划类型",
            "客户全称",
            "机构",
            "年金计划号",
            "流失日期",
            "计划规模",
            "受托人",
            "company_id",
            "区域",
            "年金中心",
            "上报人",
            "考核标签",
        ]
    )
    trustee.append(
        [
            "2026年03月",
            "受托",
            "集合",
            "共享客户（流失）",
            "北京",
            "",
            "",
            "80",
            "原受托机构A",
            "company-001",
            "华北",
            "中心A",
            "测试",
            "drop-me",
        ]
    )
    investee = workbook.create_sheet("企年投资流失(解约)")
    investee.append(
        [
            "上报月份",
            "业务类型",
            "计划类型",
            "客户全称",
            "机构",
            "年金计划号",
            "流失日期",
            "计划规模",
            "受托人",
            "company_id",
            "区域",
            "年金中心",
            "上报人",
            "考核标签",
        ]
    )
    investee.append(
        [
            "2026年03月",
            "投管",
            "单一",
            "新客流失",
            "未知机构",
            "",
            "2026-03-15",
            "60",
            "原受托机构B",
            "",
            "华东",
            "中心B",
            "测试",
            "drop-me",
        ]
    )
    workbook.save(workbook_path)

    result = AnnualLossIntakeService().read_batch(
        run_id="run-001",
        period="2026-03",
        source_files=[workbook_path],
    )

    assert result.batch.batch_id == "annual_loss:2026-03"
    assert result.batch.row_count == 2
    assert [record.anchor_row_no for record in result.records] == [2, 3]
    assert [record.origin_row_nos for record in result.records] == [[2], [2]]
    assert [record.raw_payload["source_sheet"] for record in result.records] == [
        "企年受托流失(解约)",
        "企年投资流失(解约)",
    ]
    assert [record.raw_payload["客户全称"] for record in result.records] == [
        "共享客户（流失）",
        "新客流失",
    ]


def test_annual_loss_intake_skips_trailing_empty_rows(tmp_path) -> None:
    workbook_path = tmp_path / "annual_loss_2026_04.xlsx"
    workbook = Workbook()
    trustee = workbook.active
    trustee.title = "企年受托流失(解约)"
    trustee.append(
        [
            "上报月份",
            "业务类型",
            "计划类型",
            "客户全称",
            "机构",
            "年金计划号",
            "流失日期",
            "计划规模",
            "受托人",
            "company_id",
            "区域",
            "年金中心",
            "上报人",
            "考核标签",
        ]
    )
    trustee.append(
        [
            "2026年04月",
            "受托",
            "集合",
            "客户A",
            "北京",
            "",
            "",
            "20",
            "原受托机构A",
            "company-001",
            "华北",
            "中心A",
            "测试",
            "drop-me",
        ]
    )
    trustee.append([None] * 14)

    investee = workbook.create_sheet("企年投资流失(解约)")
    investee.append(
        [
            "上报月份",
            "业务类型",
            "计划类型",
            "客户全称",
            "机构",
            "年金计划号",
            "流失日期",
            "计划规模",
            "受托人",
            "company_id",
            "区域",
            "年金中心",
            "上报人",
            "考核标签",
        ]
    )
    investee.append(
        [
            "2026年04月",
            "投管",
            "单一",
            "客户B",
            "上海",
            "",
            "2026-04-20",
            "15",
            "原受托机构B",
            "",
            "华东",
            "中心B",
            "测试",
            "drop-me",
        ]
    )
    investee.append([""] * 14)
    workbook.save(workbook_path)

    result = AnnualLossIntakeService().read_batch(
        run_id="run-002",
        period="2026-04",
        source_files=[workbook_path],
    )

    assert result.batch.row_count == 2
    assert [record.anchor_row_no for record in result.records] == [2, 3]
    assert [record.origin_row_nos for record in result.records] == [[2], [2]]
    assert [record.raw_payload["客户全称"] for record in result.records] == [
        "客户A",
        "客户B",
    ]
