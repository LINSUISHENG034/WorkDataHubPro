from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from work_data_hub_pro.apps.orchestration.replay.annual_award_slice import (
    run_annual_award_slice,
)
from work_data_hub_pro.apps.orchestration.replay.annual_loss_slice import (
    run_annual_loss_slice,
)


def _write_award_replay_assets(
    replay_root: Path,
    *,
    legacy_snapshot_rows: list[dict[str, object]],
    include_intermediate_baselines: bool = False,
) -> None:
    """Write award replay root assets."""
    replay_root.mkdir(parents=True, exist_ok=True)
    (replay_root / "annuity_performance_fixture_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "record_id": "perf-001",
                    "company_id": "company-001",
                    "plan_code": "P9001",
                    "period": "2026-03",
                    "source_record_id": "perf-001",
                }
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    (replay_root / "annual_loss_fixture_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "company_id": "company-001",
                    "plan_code": "LOSS-99",
                    "period": "2026-03",
                    "loss_code": "LOSS-99",
                    "source_sheet": "LossRegister",
                    "source_record_id": "loss-001",
                }
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    (replay_root / "customer_plan_history_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "company_id": "company-001",
                    "product_line_code": "PL-RET",
                    "plan_code": "P9001",
                    "effective_period": "2025-12",
                }
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    (replay_root / "legacy_monthly_snapshot_2026_03.json").write_text(
        json.dumps(legacy_snapshot_rows, indent=2),
        encoding="utf-8",
    )
    if include_intermediate_baselines:
        (replay_root / "legacy_reference_derivation_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "target_object": "company_reference",
                        "candidate_payload": {
                            "company_id": "company-001",
                            "company_name": "Acme",
                            "period": "2026-03",
                        },
                        "source_record_ids": ["award-001"],
                        "derivation_rule_id": "annual-award-company-reference",
                        "derivation_rule_version": "1",
                    },
                    {
                        "target_object": "customer_master_signal",
                        "candidate_payload": {
                            "company_id": "company-001",
                            "plan_code": "P9001",
                            "period": "2026-03",
                            "signal_type": "annual_award",
                        },
                        "source_record_ids": ["award-001"],
                        "derivation_rule_id": "annual-award-customer-signal",
                        "derivation_rule_version": "1",
                    },
                ],
                indent=2,
            ),
            encoding="utf-8",
        )
        (replay_root / "legacy_fact_processing_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "record_id": "award-001",
                        "company_id": "company-001",
                        "plan_code": "P9001",
                        "period": "2026-03",
                        "award_amount": 5000,
                        "source_sheet": "TrusteeAwards",
                    },
                    {
                        "record_id": "award-002",
                        "company_id": "company-002",
                        "plan_code": "S9009",
                        "period": "2026-03",
                        "award_amount": 1000,
                        "source_sheet": "InvesteeAwards",
                    },
                ],
                indent=2,
            ),
            encoding="utf-8",
        )
        (replay_root / "legacy_identity_resolution_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "record_id": "award-001",
                        "resolved_identity": "company-001",
                        "resolution_method": "static",
                        "fallback_level": "none",
                        "evidence_refs": [],
                    },
                    {
                        "record_id": "award-002",
                        "resolved_identity": "company-002",
                        "resolution_method": "static",
                        "fallback_level": "none",
                        "evidence_refs": [],
                    },
                ],
                indent=2,
            ),
            encoding="utf-8",
        )
        (replay_root / "legacy_contract_state_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "period": "2026-03",
                        "contract_state_rows": 1,
                        "award_fixture_rows": 1,
                        "loss_fixture_rows": 0,
                    }
                ],
                indent=2,
            ),
            encoding="utf-8",
        )


def _write_loss_replay_assets(
    replay_root: Path,
    *,
    legacy_snapshot_rows: list[dict[str, object]],
    include_intermediate_baselines: bool = False,
) -> None:
    """Write loss replay root assets."""
    replay_root.mkdir(parents=True, exist_ok=True)
    (replay_root / "annuity_performance_fixture_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "record_id": "perf-001",
                    "company_id": "company-001",
                    "plan_code": "P9001",
                    "period": "2026-03",
                    "source_record_id": "perf-001",
                }
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    (replay_root / "annual_award_fixture_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "company_id": "company-001",
                    "plan_code": "P9001",
                    "period": "2026-03",
                    "award_code": "AWARD-01",
                    "source_sheet": "AwardRegister",
                    "source_record_id": "award-001",
                }
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    (replay_root / "customer_plan_history_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "company_id": "company-001",
                    "product_line_code": "PL202",
                    "plan_code": "P9001",
                    "effective_period": "2025-12",
                }
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    (replay_root / "legacy_monthly_snapshot_2026_03.json").write_text(
        json.dumps(legacy_snapshot_rows, indent=2),
        encoding="utf-8",
    )
    if include_intermediate_baselines:
        (replay_root / "legacy_reference_derivation_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "target_object": "company_reference",
                        "candidate_payload": {
                            "company_id": "company-001",
                            "company_name": "共享客户（流失）",
                            "period": "2026-03",
                        },
                        "source_record_ids": ["loss-001"],
                        "derivation_rule_id": "annual-loss-company-reference",
                        "derivation_rule_version": "1",
                    },
                    {
                        "target_object": "customer_loss_signal",
                        "candidate_payload": {
                            "company_id": "company-002",
                            "plan_code": "S9009",
                            "period": "2026-03",
                            "signal_type": "annual_loss",
                        },
                        "source_record_ids": ["loss-002"],
                        "derivation_rule_id": "annual-loss-customer-signal",
                        "derivation_rule_version": "1",
                    },
                ],
                indent=2,
            ),
            encoding="utf-8",
        )
        (replay_root / "legacy_fact_processing_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "record_id": "loss-001",
                        "company_id": "company-001",
                        "plan_code": "P9001",
                        "period": "2026-03",
                        "loss_amount": 80,
                        "source_sheet": "企年受托流失(解约)",
                    },
                    {
                        "record_id": "loss-002",
                        "company_id": "company-002",
                        "plan_code": "S9009",
                        "period": "2026-03",
                        "loss_amount": 60,
                        "source_sheet": "企年投资流失(解约)",
                    },
                ],
                indent=2,
            ),
            encoding="utf-8",
        )
        (replay_root / "legacy_identity_resolution_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "record_id": "loss-001",
                        "resolved_identity": "company-001",
                        "resolution_method": "static",
                        "fallback_level": "none",
                        "evidence_refs": [],
                    },
                    {
                        "record_id": "loss-002",
                        "resolved_identity": "company-002",
                        "resolution_method": "static",
                        "fallback_level": "none",
                        "evidence_refs": [],
                    },
                ],
                indent=2,
            ),
            encoding="utf-8",
        )
        (replay_root / "legacy_contract_state_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "period": "2026-03",
                        "contract_state_rows": 1,
                        "award_fixture_rows": 1,
                        "loss_fixture_rows": 1,
                    }
                ],
                indent=2,
            ),
            encoding="utf-8",
        )


def _write_award_workbook(path: Path) -> None:
    workbook = Workbook()
    trustee = workbook.active
    trustee.title = "TrusteeAwards"
    trustee.append(
        [
            "company_name",
            "source_company_id",
            "plan_code",
            "plan_type",
            "product_line_code",
            "period",
            "award_amount",
        ]
    )
    trustee.append(["Acme", "company-001", "", "collective", "pl-ret", "2026-03", "5000"])
    investee = workbook.create_sheet("InvesteeAwards")
    investee.append(
        [
            "company_name",
            "source_company_id",
            "plan_code",
            "plan_type",
            "product_line_code",
            "period",
            "award_amount",
        ]
    )
    investee.append(["Beta", "", "", "single", "pl-alt", "2026-03", "1000"])
    workbook.save(path)


def _write_loss_workbook(path: Path) -> None:
    workbook = Workbook()
    trustee = workbook.active
    trustee.title = "企年受托流失(解约)"
    trustee.append(
        [
            "上报月份",
            "业务类型",
            "计划类型",
            "客户全称",
            "机构",
            "年金计划号",
            "流失日期",
            "计划规模",
            "受托人",
            "company_id",
        ]
    )
    trustee.append(
        [
            "2026年03月",
            "受托",
            "集合",
            "共享客户（流失）",
            "北京",
            "",
            "",
            "80",
            "原受托机构A",
            "company-001",
        ]
    )
    investee = workbook.create_sheet("企年投资流失(解约)")
    investee.append(
        [
            "上报月份",
            "业务类型",
            "计划类型",
            "客户全称",
            "机构",
            "年金计划号",
            "流失日期",
            "计划规模",
            "受托人",
            "company_id",
        ]
    )
    investee.append(
        [
            "2026年03月",
            "投管",
            "单一",
            "新客流失",
            "未知机构",
            "",
            "2026-03-15",
            "60",
            "原受托机构B",
            "",
        ]
    )
    workbook.save(path)


# ---------------------------------------------------------------------------
# Tests for truthful intermediate gates
# ---------------------------------------------------------------------------


def test_event_domain_gates_use_same_checkpoint_names(tmp_path) -> None:
    """Existing test: award and loss slices produce the same checkpoint names."""
    award_workbook = tmp_path / "annual_award.xlsx"
    loss_workbook = tmp_path / "annual_loss.xlsx"
    _write_award_workbook(award_workbook)
    _write_loss_workbook(loss_workbook)
    award_replay_root = tmp_path / "reference" / "historical_replays" / "annual_award"
    loss_replay_root = tmp_path / "reference" / "historical_replays" / "annual_loss"

    _write_award_replay_assets(
        award_replay_root,
        legacy_snapshot_rows=[
            {
                "period": "2026-03",
                "contract_state_rows": 1,
                "award_fixture_rows": 1,
                "loss_fixture_rows": 0,
            }
        ],
        include_intermediate_baselines=True,
    )
    _write_loss_replay_assets(
        loss_replay_root,
        legacy_snapshot_rows=[
            {
                "period": "2026-03",
                "contract_state_rows": 1,
                "award_fixture_rows": 1,
                "loss_fixture_rows": 1,
            }
        ],
        include_intermediate_baselines=True,
    )

    award_outcome = run_annual_award_slice(
        workbook=award_workbook,
        period="2026-03",
        replay_root=award_replay_root,
    )
    loss_outcome = run_annual_loss_slice(
        workbook=loss_workbook,
        period="2026-03",
        replay_root=loss_replay_root,
    )

    expected_checkpoints = [
        "source_intake",
        "fact_processing",
        "identity_resolution",
        "reference_derivation",
        "contract_state",
        "monthly_snapshot",
    ]

    assert [result.checkpoint_name for result in award_outcome.checkpoint_results] == expected_checkpoints
    assert [result.checkpoint_name for result in loss_outcome.checkpoint_results] == expected_checkpoints


def test_event_domain_intermediate_checkpoint_uses_baseline_award(tmp_path) -> None:
    """Test 2 (TDD RED): award fact_processing, identity_resolution, and contract_state
    each fail independently when the corresponding baseline file is edited to a mismatching value."""
    workbook_path = tmp_path / "annual_award.xlsx"
    _write_award_workbook(workbook_path)
    replay_root = tmp_path / "reference" / "historical_replays" / "annual_award"

    _write_award_replay_assets(
        replay_root,
        legacy_snapshot_rows=[
            {
                "period": "2026-03",
                "contract_state_rows": 1,
                "award_fixture_rows": 1,
                "loss_fixture_rows": 0,
            }
        ],
        include_intermediate_baselines=True,
    )

    # --- Step 1: Verify passing with matching baselines ---
    outcome_pass = run_annual_award_slice(
        workbook=workbook_path,
        period="2026-03",
        replay_root=replay_root,
    )
    for ckpt in ("fact_processing", "identity_resolution", "contract_state"):
        result = next(r for r in outcome_pass.checkpoint_results if r.checkpoint_name == ckpt)
        assert result.status == "passed", f"{ckpt} should pass with matching baseline"

    # --- Step 2: fact_processing can fail independently ---
    (replay_root / "legacy_fact_processing_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "record_id": "award-001",
                    "company_id": "company-999",  # Deliberately wrong
                    "plan_code": "WRONG",
                    "period": "2026-03",
                    "award_amount": 0,
                    "source_sheet": "TrusteeAwards",
                },
                {
                    "record_id": "award-002",
                    "company_id": "company-999",
                    "plan_code": "WRONG",
                    "period": "2026-03",
                    "award_amount": 0,
                    "source_sheet": "InvesteeAwards",
                },
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    outcome_fp_fail = run_annual_award_slice(
        workbook=workbook_path,
        period="2026-03",
        replay_root=replay_root,
    )
    fp_result = next(
        r for r in outcome_fp_fail.checkpoint_results if r.checkpoint_name == "fact_processing"
    )
    assert fp_result.status == "failed", (
        "fact_processing should fail with mismatched baseline"
    )
    # identity_resolution and contract_state still pass (independent failure)
    for ckpt in ("identity_resolution", "contract_state"):
        result = next(r for r in outcome_fp_fail.checkpoint_results if r.checkpoint_name == ckpt)
        assert result.status == "passed", f"{ckpt} should still pass"

    # Restore fact_processing baseline
    (replay_root / "legacy_fact_processing_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "record_id": "award-001",
                    "company_id": "company-001",
                    "plan_code": "P9001",
                    "period": "2026-03",
                    "award_amount": 5000,
                    "source_sheet": "TrusteeAwards",
                },
                {
                    "record_id": "award-002",
                    "company_id": "company-002",
                    "plan_code": "S9009",
                    "period": "2026-03",
                    "award_amount": 1000,
                    "source_sheet": "InvesteeAwards",
                },
            ],
            indent=2,
        ),
        encoding="utf-8",
    )

    # --- Step 3: identity_resolution can fail independently ---
    (replay_root / "legacy_identity_resolution_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "record_id": "award-001",
                    "resolved_identity": "company-999",  # Deliberately wrong
                    "resolution_method": "static",
                    "fallback_level": "none",
                    "evidence_refs": [],
                },
                {
                    "record_id": "award-002",
                    "resolved_identity": "company-999",
                    "resolution_method": "static",
                    "fallback_level": "none",
                    "evidence_refs": [],
                },
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    outcome_ir_fail = run_annual_award_slice(
        workbook=workbook_path,
        period="2026-03",
        replay_root=replay_root,
    )
    ir_result = next(
        r for r in outcome_ir_fail.checkpoint_results if r.checkpoint_name == "identity_resolution"
    )
    assert ir_result.status == "failed", (
        "identity_resolution should fail with mismatched baseline"
    )

    # Restore identity_resolution baseline
    (replay_root / "legacy_identity_resolution_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "record_id": "award-001",
                    "resolved_identity": "company-001",
                    "resolution_method": "static",
                    "fallback_level": "none",
                    "evidence_refs": [],
                },
                {
                    "record_id": "award-002",
                    "resolved_identity": "company-002",
                    "resolution_method": "static",
                    "fallback_level": "none",
                    "evidence_refs": [],
                },
            ],
            indent=2,
        ),
        encoding="utf-8",
    )

    # --- Step 4: contract_state can fail independently ---
    (replay_root / "legacy_contract_state_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "period": "2026-03",
                    "contract_state_rows": 99,  # Deliberately wrong
                    "award_fixture_rows": 99,
                    "loss_fixture_rows": 99,
                }
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    outcome_cs_fail = run_annual_award_slice(
        workbook=workbook_path,
        period="2026-03",
        replay_root=replay_root,
    )
    cs_result = next(
        r for r in outcome_cs_fail.checkpoint_results if r.checkpoint_name == "contract_state"
    )
    assert cs_result.status == "failed", (
        "contract_state should fail with mismatched baseline"
    )


def test_event_domain_intermediate_checkpoint_uses_baseline_loss(tmp_path) -> None:
    """Test 2 (TDD RED): loss fact_processing, identity_resolution, and contract_state
    each fail independently when the corresponding baseline file is edited."""
    workbook_path = tmp_path / "annual_loss.xlsx"
    _write_loss_workbook(workbook_path)
    replay_root = tmp_path / "reference" / "historical_replays" / "annual_loss"

    _write_loss_replay_assets(
        replay_root,
        legacy_snapshot_rows=[
            {
                "period": "2026-03",
                "contract_state_rows": 1,
                "award_fixture_rows": 1,
                "loss_fixture_rows": 1,
            }
        ],
        include_intermediate_baselines=True,
    )

    # --- Step 1: Verify passing with matching baselines ---
    outcome_pass = run_annual_loss_slice(
        workbook=workbook_path,
        period="2026-03",
        replay_root=replay_root,
    )
    for ckpt in ("fact_processing", "identity_resolution", "contract_state"):
        result = next(r for r in outcome_pass.checkpoint_results if r.checkpoint_name == ckpt)
        assert result.status == "passed", f"{ckpt} should pass with matching baseline"

    # --- Step 2: fact_processing fails independently ---
    (replay_root / "legacy_fact_processing_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "record_id": "loss-001",
                    "company_id": "company-999",  # Wrong
                    "plan_code": "WRONG",
                    "period": "2026-03",
                    "loss_amount": 0,
                    "source_sheet": "企年受托流失(解约)",
                },
                {
                    "record_id": "loss-002",
                    "company_id": "company-999",
                    "plan_code": "WRONG",
                    "period": "2026-03",
                    "loss_amount": 0,
                    "source_sheet": "企年投资流失(解约)",
                },
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    outcome_fp_fail = run_annual_loss_slice(
        workbook=workbook_path,
        period="2026-03",
        replay_root=replay_root,
    )
    fp_result = next(
        r for r in outcome_fp_fail.checkpoint_results if r.checkpoint_name == "fact_processing"
    )
    assert fp_result.status == "failed", (
        "fact_processing should fail with mismatched baseline"
    )
    for ckpt in ("identity_resolution", "contract_state"):
        result = next(r for r in outcome_fp_fail.checkpoint_results if r.checkpoint_name == ckpt)
        assert result.status == "passed", f"{ckpt} should still pass"

    # Restore
    (replay_root / "legacy_fact_processing_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "record_id": "loss-001",
                    "company_id": "company-001",
                    "plan_code": "P9001",
                    "period": "2026-03",
                    "loss_amount": 80,
                    "source_sheet": "企年受托流失(解约)",
                },
                {
                    "record_id": "loss-002",
                    "company_id": "company-002",
                    "plan_code": "S9009",
                    "period": "2026-03",
                    "loss_amount": 60,
                    "source_sheet": "企年投资流失(解约)",
                },
            ],
            indent=2,
        ),
        encoding="utf-8",
    )

    # --- Step 3: contract_state fails independently ---
    (replay_root / "legacy_contract_state_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "period": "2026-03",
                    "contract_state_rows": 99,
                    "award_fixture_rows": 99,
                    "loss_fixture_rows": 99,
                }
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    outcome_cs_fail = run_annual_loss_slice(
        workbook=workbook_path,
        period="2026-03",
        replay_root=replay_root,
    )
    cs_result = next(
        r for r in outcome_cs_fail.checkpoint_results if r.checkpoint_name == "contract_state"
    )
    assert cs_result.status == "failed", (
        "contract_state should fail with mismatched baseline"
    )


def test_event_domain_failed_runs_write_same_package(tmp_path) -> None:
    """Existing test: failed runs write the same package structure."""
    award_workbook = tmp_path / "annual_award.xlsx"
    loss_workbook = tmp_path / "annual_loss.xlsx"
    _write_award_workbook(award_workbook)
    _write_loss_workbook(loss_workbook)
    award_replay_root = tmp_path / "reference" / "historical_replays" / "annual_award"
    loss_replay_root = tmp_path / "reference" / "historical_replays" / "annual_loss"

    # Use intermediate baselines with mismatching monthly_snapshot
    _write_award_replay_assets(award_replay_root, legacy_snapshot_rows=[{"period": "2026-03", "contract_state_rows": 99, "award_fixture_rows": 99, "loss_fixture_rows": 99}], include_intermediate_baselines=True)
    _write_loss_replay_assets(loss_replay_root, legacy_snapshot_rows=[{"period": "2026-03", "contract_state_rows": 99, "award_fixture_rows": 99, "loss_fixture_rows": 99}], include_intermediate_baselines=True)

    award_outcome = run_annual_award_slice(
        workbook=award_workbook,
        period="2026-03",
        replay_root=award_replay_root,
    )
    loss_outcome = run_annual_loss_slice(
        workbook=loss_workbook,
        period="2026-03",
        replay_root=loss_replay_root,
    )

    award_package_root = (
        award_replay_root / "evidence" / "comparison_runs" / award_outcome.comparison_run_id
    )
    loss_package_root = (
        loss_replay_root / "evidence" / "comparison_runs" / loss_outcome.comparison_run_id
    )

    for package_root in (award_package_root, loss_package_root):
        assert (package_root / "manifest.json").exists()
        assert (package_root / "gate-summary.json").exists()
        assert (package_root / "checkpoint-results.json").exists()
        assert (package_root / "source-intake-adaptation.json").exists()
        assert (package_root / "lineage-impact.json").exists()
        assert (package_root / "publication-results.json").exists()
        assert (package_root / "compatibility-case.json").exists()
        assert (package_root / "report.md").exists()
        assert (package_root / "diffs" / "monthly_snapshot.json").exists()
