from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from work_data_hub_pro.apps.orchestration.replay.annuity_performance_slice import (
    run_annuity_performance_slice,
)
from work_data_hub_pro.apps.orchestration.replay.annual_award_slice import (
    run_annual_award_slice,
)
from work_data_hub_pro.apps.orchestration.replay.annual_loss_slice import (
    run_annual_loss_slice,
)
from work_data_hub_pro.apps.orchestration.replay.errors import (
    ReplayAssetNotFoundError,
)


def write_annuity_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "AnnuityPerformance"
    sheet.append(["company_name", "plan_code", "period", "sales_amount"])
    sheet.append(["Acme", "PLAN-A", "2026-03", "1200.50"])
    workbook.save(path)


def _write_annuity_replay_assets(
    replay_root: Path,
    *,
    legacy_snapshot_rows: list[dict[str, object]],
    include_intermediate_baselines: bool = False,
) -> None:
    """Write annuity replay root assets."""
    replay_root.mkdir(parents=True, exist_ok=True)
    (replay_root / "annual_award_fixture_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "company_id": "company-001",
                    "plan_code": "PLAN-A",
                    "period": "2026-03",
                    "award_code": "AWARD-01",
                    "source_sheet": "AwardRegister",
                    "source_record_id": "award-001",
                }
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    (replay_root / "annual_loss_fixture_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "company_id": "company-001",
                    "plan_code": "LOSS-99",
                    "period": "2026-03",
                    "loss_code": "LOSS-99",
                    "source_sheet": "LossRegister",
                    "source_record_id": "loss-001",
                }
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    (replay_root / "legacy_monthly_snapshot_2026_03.json").write_text(
        json.dumps(legacy_snapshot_rows, indent=2),
        encoding="utf-8",
    )
    if include_intermediate_baselines:
        (replay_root / "legacy_reference_derivation_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "target_object": "company_reference",
                        "candidate_payload": {
                            "company_id": "company-001",
                            "company_name": "Acme",
                            "period": "2026-03",
                        },
                        "source_record_ids": ["perf-001"],
                        "derivation_rule_id": "annuity-performance-company-reference",
                        "derivation_rule_version": "1",
                    }
                ],
                indent=2,
            ),
            encoding="utf-8",
        )
        (replay_root / "legacy_fact_processing_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "record_id": "perf-001",
                        "company_id": "company-001",
                        "plan_code": "PLAN-A",
                        "period": "2026-03",
                        "ending_assets": 1200.5,
                    }
                ],
                indent=2,
            ),
            encoding="utf-8",
        )
        (replay_root / "legacy_identity_resolution_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "record_id": "perf-001",
                        "resolved_identity": "company-001",
                        "resolution_method": "static",
                        "fallback_level": "none",
                        "evidence_refs": [],
                    }
                ],
                indent=2,
            ),
            encoding="utf-8",
        )
        (replay_root / "legacy_contract_state_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "period": "2026-03",
                        "contract_state_rows": 1,
                        "award_fixture_rows": 1,
                        "loss_fixture_rows": 0,
                    }
                ],
                indent=2,
            ),
            encoding="utf-8",
        )


def write_award_workbook(path: Path) -> None:
    workbook = Workbook()
    trustee = workbook.active
    trustee.title = "TrusteeAwards"
    trustee.append(
        [
            "company_name",
            "source_company_id",
            "plan_code",
            "plan_type",
            "product_line_code",
            "period",
            "award_amount",
        ]
    )
    trustee.append(["Acme", "company-001", "", "collective", "pl-ret", "2026-03", "5000"])
    investee = workbook.create_sheet("InvesteeAwards")
    investee.append(
        [
            "company_name",
            "source_company_id",
            "plan_code",
            "plan_type",
            "product_line_code",
            "period",
            "award_amount",
        ]
    )
    investee.append(["Beta", "", "", "single", "pl-alt", "2026-03", "1000"])
    workbook.save(path)


def _write_award_replay_assets(
    replay_root: Path,
    *,
    legacy_snapshot_rows: list[dict[str, object]],
    include_intermediate_baselines: bool = False,
) -> None:
    """Write award replay root assets."""
    replay_root.mkdir(parents=True, exist_ok=True)
    (replay_root / "annuity_performance_fixture_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "record_id": "perf-001",
                    "company_id": "company-001",
                    "plan_code": "P9001",
                    "period": "2026-03",
                    "source_record_id": "perf-001",
                }
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    (replay_root / "annual_loss_fixture_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "company_id": "company-001",
                    "plan_code": "LOSS-99",
                    "period": "2026-03",
                    "loss_code": "LOSS-99",
                    "source_sheet": "LossRegister",
                    "source_record_id": "loss-001",
                }
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    (replay_root / "customer_plan_history_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "company_id": "company-001",
                    "product_line_code": "PL-RET",
                    "plan_code": "P9001",
                    "effective_period": "2025-12",
                }
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    (replay_root / "legacy_monthly_snapshot_2026_03.json").write_text(
        json.dumps(legacy_snapshot_rows, indent=2),
        encoding="utf-8",
    )
    if include_intermediate_baselines:
        (replay_root / "legacy_reference_derivation_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "target_object": "company_reference",
                        "candidate_payload": {
                            "company_id": "company-001",
                            "company_name": "Acme",
                            "period": "2026-03",
                        },
                        "source_record_ids": ["award-001"],
                        "derivation_rule_id": "annual-award-company-reference",
                        "derivation_rule_version": "1",
                    },
                    {
                        "target_object": "customer_master_signal",
                        "candidate_payload": {
                            "company_id": "company-001",
                            "plan_code": "P9001",
                            "period": "2026-03",
                            "signal_type": "annual_award",
                        },
                        "source_record_ids": ["award-001"],
                        "derivation_rule_id": "annual-award-customer-signal",
                        "derivation_rule_version": "1",
                    },
                ],
                indent=2,
            ),
            encoding="utf-8",
        )
        (replay_root / "legacy_fact_processing_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "record_id": "award-001",
                        "company_id": "company-001",
                        "plan_code": "P9001",
                        "period": "2026-03",
                        "award_amount": 5000,
                        "source_sheet": "TrusteeAwards",
                    },
                    {
                        "record_id": "award-002",
                        "company_id": "company-002",
                        "plan_code": "S9009",
                        "period": "2026-03",
                        "award_amount": 1000,
                        "source_sheet": "InvesteeAwards",
                    },
                ],
                indent=2,
            ),
            encoding="utf-8",
        )
        (replay_root / "legacy_identity_resolution_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "record_id": "award-001",
                        "resolved_identity": "company-001",
                        "resolution_method": "static",
                        "fallback_level": "none",
                        "evidence_refs": [],
                    },
                    {
                        "record_id": "award-002",
                        "resolved_identity": "company-002",
                        "resolution_method": "static",
                        "fallback_level": "none",
                        "evidence_refs": [],
                    },
                ],
                indent=2,
            ),
            encoding="utf-8",
        )
        (replay_root / "legacy_contract_state_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "period": "2026-03",
                        "contract_state_rows": 1,
                        "award_fixture_rows": 1,
                        "loss_fixture_rows": 0,
                    }
                ],
                indent=2,
            ),
            encoding="utf-8",
        )


def write_loss_workbook(path: Path) -> None:
    workbook = Workbook()
    trustee = workbook.active
    trustee.title = "企年受托流失(解约)"
    trustee.append(
        [
            "上报月份",
            "业务类型",
            "计划类型",
            "客户全称",
            "机构",
            "年金计划号",
            "流失日期",
            "计划规模",
            "受托人",
            "company_id",
        ]
    )
    trustee.append(
        [
            "2026年03月",
            "受托",
            "集合",
            "共享客户（流失）",
            "北京",
            "",
            "",
            "80",
            "原受托机构A",
            "company-001",
        ]
    )
    investee = workbook.create_sheet("企年投资流失(解约)")
    investee.append(
        [
            "上报月份",
            "业务类型",
            "计划类型",
            "客户全称",
            "机构",
            "年金计划号",
            "流失日期",
            "计划规模",
            "受托人",
            "company_id",
        ]
    )
    investee.append(
        [
            "2026年03月",
            "投管",
            "单一",
            "新客流失",
            "未知机构",
            "",
            "2026-03-15",
            "60",
            "原受托机构B",
            "",
        ]
    )
    workbook.save(path)


def _bootstrap_intermediate_baselines(runner, workbook_path: Path, replay_root: Path) -> dict[str, list[dict[str, object]]]:
    for checkpoint_name in (
        "reference_derivation",
        "fact_processing",
        "identity_resolution",
        "contract_state",
    ):
        (replay_root / f"legacy_{checkpoint_name}_2026_03.json").write_text(
            "[]",
            encoding="utf-8",
        )
    outcome = runner(
        workbook=workbook_path,
        period="2026-03",
        replay_root=replay_root,
    )
    for checkpoint_name in (
        "reference_derivation",
        "fact_processing",
        "identity_resolution",
        "contract_state",
    ):
        (replay_root / f"legacy_{checkpoint_name}_2026_03.json").write_text(
            json.dumps(
                outcome.intermediate_payloads[checkpoint_name],
                indent=2,
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
    return outcome.intermediate_payloads


def _write_loss_replay_assets(
    replay_root: Path,
    *,
    legacy_snapshot_rows: list[dict[str, object]],
    include_intermediate_baselines: bool = False,
) -> None:
    """Write loss replay root assets."""
    replay_root.mkdir(parents=True, exist_ok=True)
    (replay_root / "annuity_performance_fixture_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "record_id": "perf-001",
                    "company_id": "company-001",
                    "plan_code": "P9001",
                    "period": "2026-03",
                    "source_record_id": "perf-001",
                }
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    (replay_root / "annual_award_fixture_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "company_id": "company-001",
                    "plan_code": "P9001",
                    "period": "2026-03",
                    "award_code": "AWARD-01",
                    "source_sheet": "AwardRegister",
                    "source_record_id": "award-001",
                }
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    (replay_root / "customer_plan_history_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "company_id": "company-001",
                    "product_line_code": "PL202",
                    "plan_code": "P9001",
                    "effective_period": "2025-12",
                }
            ],
            indent=2,
        ),
        encoding="utf-8",
    )
    (replay_root / "legacy_monthly_snapshot_2026_03.json").write_text(
        json.dumps(legacy_snapshot_rows, indent=2),
        encoding="utf-8",
    )
    if include_intermediate_baselines:
        (replay_root / "legacy_reference_derivation_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "target_object": "company_reference",
                        "candidate_payload": {
                            "company_id": "company-001",
                            "company_name": "共享客户（流失）",
                            "period": "2026-03",
                        },
                        "source_record_ids": ["loss-001"],
                        "derivation_rule_id": "annual-loss-company-reference",
                        "derivation_rule_version": "1",
                    },
                    {
                        "target_object": "customer_loss_signal",
                        "candidate_payload": {
                            "company_id": "company-002",
                            "plan_code": "S9009",
                            "period": "2026-03",
                            "signal_type": "annual_loss",
                        },
                        "source_record_ids": ["loss-002"],
                        "derivation_rule_id": "annual-loss-customer-signal",
                        "derivation_rule_version": "1",
                    },
                ],
                indent=2,
            ),
            encoding="utf-8",
        )
        (replay_root / "legacy_fact_processing_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "record_id": "loss-001",
                        "company_id": "company-001",
                        "plan_code": "P9001",
                        "period": "2026-03",
                        "loss_amount": 80,
                        "source_sheet": "企年受托流失(解约)",
                    },
                    {
                        "record_id": "loss-002",
                        "company_id": "company-002",
                        "plan_code": "S9009",
                        "period": "2026-03",
                        "loss_amount": 60,
                        "source_sheet": "企年投资流失(解约)",
                    },
                ],
                indent=2,
            ),
            encoding="utf-8",
        )
        (replay_root / "legacy_identity_resolution_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "record_id": "loss-001",
                        "resolved_identity": "company-001",
                        "resolution_method": "static",
                        "fallback_level": "none",
                        "evidence_refs": [],
                    },
                    {
                        "record_id": "loss-002",
                        "resolved_identity": "company-002",
                        "resolution_method": "static",
                        "fallback_level": "none",
                        "evidence_refs": [],
                    },
                ],
                indent=2,
            ),
            encoding="utf-8",
        )
        (replay_root / "legacy_contract_state_2026_03.json").write_text(
            json.dumps(
                [
                    {
                        "period": "2026-03",
                        "contract_state_rows": 1,
                        "award_fixture_rows": 1,
                        "loss_fixture_rows": 1,
                    }
                ],
                indent=2,
            ),
            encoding="utf-8",
        )


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_reference_derivation_checkpoint_present(tmp_path) -> None:
    """reference_derivation is present in checkpoint list (existing behavior)."""
    annuity_workbook = tmp_path / "annuity.xlsx"
    award_workbook = tmp_path / "annual_award.xlsx"
    loss_workbook = tmp_path / "annual_loss.xlsx"
    write_annuity_workbook(annuity_workbook)
    write_award_workbook(award_workbook)
    write_loss_workbook(loss_workbook)

    annuity_root = tmp_path / "reference" / "historical_replays" / "annuity_performance"
    award_root = tmp_path / "reference" / "historical_replays" / "annual_award"
    loss_root = tmp_path / "reference" / "historical_replays" / "annual_loss"

    _write_annuity_replay_assets(
        annuity_root,
        legacy_snapshot_rows=[{"period": "2026-03", "contract_state_rows": 1, "award_fixture_rows": 1, "loss_fixture_rows": 0}],
    )
    _bootstrap_intermediate_baselines(
        run_annuity_performance_slice,
        annuity_workbook,
        annuity_root,
    )
    _write_award_replay_assets(
        award_root,
        legacy_snapshot_rows=[{"period": "2026-03", "contract_state_rows": 1, "award_fixture_rows": 1, "loss_fixture_rows": 0}],
    )
    _bootstrap_intermediate_baselines(
        run_annual_award_slice,
        award_workbook,
        award_root,
    )
    _write_loss_replay_assets(
        loss_root,
        legacy_snapshot_rows=[{"period": "2026-03", "contract_state_rows": 1, "award_fixture_rows": 1, "loss_fixture_rows": 1}],
    )
    _bootstrap_intermediate_baselines(
        run_annual_loss_slice,
        loss_workbook,
        loss_root,
    )

    annuity_outcome = run_annuity_performance_slice(
        workbook=annuity_workbook, period="2026-03", replay_root=annuity_root,
    )
    award_outcome = run_annual_award_slice(
        workbook=award_workbook, period="2026-03", replay_root=award_root,
    )
    loss_outcome = run_annual_loss_slice(
        workbook=loss_workbook, period="2026-03", replay_root=loss_root,
    )

    for outcome in (annuity_outcome, award_outcome, loss_outcome):
        assert "reference_derivation" in [
            r.checkpoint_name for r in outcome.checkpoint_results
        ]


def test_reference_derivation_requires_baseline_annuity(tmp_path) -> None:
    """Missing reference_derivation baseline raises a typed setup error."""
    workbook_path = tmp_path / "annuity.xlsx"
    write_annuity_workbook(workbook_path)
    replay_root = tmp_path / "reference" / "historical_replays" / "annuity_performance"

    _write_annuity_replay_assets(
        replay_root,
        legacy_snapshot_rows=[{"period": "2026-03", "contract_state_rows": 1, "award_fixture_rows": 1, "loss_fixture_rows": 0}],
    )
    _bootstrap_intermediate_baselines(
        run_annuity_performance_slice,
        workbook_path,
        replay_root,
    )
    # Remove reference_derivation baseline to trigger fail-closed
    (replay_root / "legacy_reference_derivation_2026_03.json").unlink(missing_ok=True)

    with pytest.raises(ReplayAssetNotFoundError, match="reference_derivation") as excinfo:
        run_annuity_performance_slice(
            workbook=workbook_path, period="2026-03", replay_root=replay_root,
        )
    assert excinfo.value.stage == "baseline_load"
    assert excinfo.value.context["checkpoint_name"] == "reference_derivation"


def test_reference_derivation_requires_baseline_award(tmp_path) -> None:
    """Missing reference_derivation baseline raises a typed setup error for award."""
    workbook_path = tmp_path / "annual_award.xlsx"
    write_award_workbook(workbook_path)
    replay_root = tmp_path / "reference" / "historical_replays" / "annual_award"

    _write_award_replay_assets(
        replay_root,
        legacy_snapshot_rows=[{"period": "2026-03", "contract_state_rows": 1, "award_fixture_rows": 1, "loss_fixture_rows": 0}],
    )
    _bootstrap_intermediate_baselines(
        run_annual_award_slice,
        workbook_path,
        replay_root,
    )
    (replay_root / "legacy_reference_derivation_2026_03.json").unlink(missing_ok=True)

    with pytest.raises(ReplayAssetNotFoundError, match="reference_derivation") as excinfo:
        run_annual_award_slice(
            workbook=workbook_path, period="2026-03", replay_root=replay_root,
        )
    assert excinfo.value.stage == "baseline_load"
    assert excinfo.value.context["checkpoint_name"] == "reference_derivation"


def test_reference_derivation_requires_baseline_loss(tmp_path) -> None:
    """Missing reference_derivation baseline raises a typed setup error for loss."""
    workbook_path = tmp_path / "annual_loss.xlsx"
    write_loss_workbook(workbook_path)
    replay_root = tmp_path / "reference" / "historical_replays" / "annual_loss"

    _write_loss_replay_assets(
        replay_root,
        legacy_snapshot_rows=[{"period": "2026-03", "contract_state_rows": 1, "award_fixture_rows": 1, "loss_fixture_rows": 1}],
    )
    _bootstrap_intermediate_baselines(
        run_annual_loss_slice,
        workbook_path,
        replay_root,
    )
    (replay_root / "legacy_reference_derivation_2026_03.json").unlink(missing_ok=True)

    with pytest.raises(ReplayAssetNotFoundError, match="reference_derivation") as excinfo:
        run_annual_loss_slice(
            workbook=workbook_path, period="2026-03", replay_root=replay_root,
        )
    assert excinfo.value.stage == "baseline_load"
    assert excinfo.value.context["checkpoint_name"] == "reference_derivation"


def test_reference_derivation_failed_run_writes_diff(tmp_path) -> None:
    """Failed reference_derivation run writes compatibility diff (existing behavior)."""
    workbook_path = tmp_path / "annuity.xlsx"
    write_annuity_workbook(workbook_path)
    replay_root = tmp_path / "reference" / "historical_replays" / "annuity_performance"

    _write_annuity_replay_assets(
        replay_root,
        legacy_snapshot_rows=[{"period": "2026-03", "contract_state_rows": 1, "award_fixture_rows": 1, "loss_fixture_rows": 0}],
    )
    _bootstrap_intermediate_baselines(
        run_annuity_performance_slice,
        workbook_path,
        replay_root,
    )
    # Overwrite reference_derivation baseline with mismatching value
    (replay_root / "legacy_reference_derivation_2026_03.json").write_text(
        json.dumps(
            [
                {
                    "target_object": "company_reference",
                    "candidate_payload": {
                        "company_id": "company-999",
                        "company_name": "Mismatch",
                        "period": "2026-03",
                    },
                    "source_record_ids": ["legacy-record"],
                    "derivation_rule_id": "legacy-derivation",
                    "derivation_rule_version": "1",
                }
            ],
            indent=2,
        ),
        encoding="utf-8",
    )

    outcome = run_annuity_performance_slice(
        workbook=workbook_path, period="2026-03", replay_root=replay_root,
    )

    package_root = replay_root / "evidence" / "comparison_runs" / outcome.comparison_run_id
    assert outcome.compatibility_case is not None
    assert outcome.compatibility_case.checkpoint_name == "reference_derivation"
    assert (package_root / "diffs" / "reference_derivation.json").exists()
